from datetime import date,timedelta
from includes.modules import types, FSMContext,json,requests,BeautifulSoup,State,StatesGroup
from includes.keyboards import Keyboards,Analytics_Keyboard
from includes.sqlite import Database
from includes.functions import Functions
from includes.states import ChannelsEdit,vip_signal_state
from aiogram.types import InputFile
import re,random,os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment

db=Database("your_database.db")
keb = Keyboards()
anal = Analytics_Keyboard()
with open('config.json', 'r', encoding='utf8') as config:
    config = json.load(config)

def write_config(config):
    with open('config.json', 'w') as file:
        json.dump(config, file, indent=4)

class analytics:
    def __init__(self, dp,bot):
        self.dp = dp
        self.bot = bot
        self.fun = Functions(self.dp, self.bot)
    def handle(self):
        @self.dp.message_handler(lambda m:m.text == "بەشداربوانی گرووپی تایبەت")
        async def vip_signal_members(message : types.Message):
            await message.answer("تكایە یەكێك لە هەڵبژاردەكان دیاری بكە :",reply_markup=anal.vip_signal_member())
        @self.dp.message_handler(lambda m:m.text == "داتاكان 📋")
        async def cmd_datas(message : types.Message):
            await message.answer("داتاكان",reply_markup=keb.get_analaytics_buttons())
        @self.dp.message_handler(lambda m:m.text in ["سەرجەم بەشداربووەكان", "سەرجەم چاوەڕوانی پارەدان","سەرجەم بەردامبووەكان","سەرجەم بەسەرچووەكان","گەڕانەوە بۆ داتاكان"])
        async def vip_signals_buttons(message:types.Message):
            members = db.get_vip_signals()
            if message.text == "سەرجەم بەشداربووەكان":
                header = ["ز", "ئایدی", "ڕۆژی بەشداری كردن", "ڕۆژی بەسەرچوون", "جۆری بەشداری كردن", "دۆخ", "نرخ", "جۆری پارەدان","#"]
                running = [header]
                expired = [header]
                pedding = [header]

                for i, member in enumerate(members, start=1):
                    entry = [i, member[1], member[2], member[3], member[4], member[5], member[6], member[7], member[8]]

                    if member[5] == "done":
                        running.append(entry)
                    elif member[5] == "pedding_payment":
                        pedding.append(entry)
                    elif member[5] == "expired":
                        expired.append(entry)
                await message.answer(f"""
زانیاری سەرجەم بەشداربووانی گرووپی تایبەت

سەرجەم بەشداربووەكان : {len(members)}
سەرجەم بەردەوامبووەكان : {len(running)}
سەرجەم بەسەرچووەكان : {len(expired)}
سەرجەم چاوەڕوانی پارەدان : {len(pedding)}

{date.today()}

""",parse_mode=types.ParseMode.MARKDOWN)
                def apply_formatting(sheet):
                    header_row = sheet[1]
                    for cell in header_row:
                        cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                def set_column_width(sheet, width):
                    for col in sheet.columns:
                        sheet.column_dimensions[col[0].column_letter].width = width

                def create_and_save_sheet(name, data, filename):
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title = name
                    for row in data:
                        sheet.append(row)
                    apply_formatting(sheet)
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                        for cell in row:
                            cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    sheet.sheet_view.rightToLeft = True
                    set_column_width(sheet, 18)
                    wb.save(filename)


                create_and_save_sheet("Running", running, "داتای بەردەوامبووان.xlsx")
                create_and_save_sheet("Expired", expired, "داتای بەسەرچووەكان.xlsx")
                create_and_save_sheet("Pending", pedding, "داتای چاوەڕوانی پارەدان.xlsx")

                with open("داتای بەردەوامبووان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
                with open("داتای بەسەرچووەكان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
                with open("داتای چاوەڕوانی پارەدان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
            if message.text == "سەرجەم بەردامبووەكان":
                header = ["ز", "ئایدی", "ڕۆژی بەشداری كردن", "ڕۆژی بەسەرچوون", "جۆری بەشداری كردن", "دۆخ", "نرخ", "جۆری پارەدان","#"]
                running = [header]
                for i, member in enumerate(members, start=1):
                    entry = [i, member[1], member[2], member[3], member[4], member[5], member[6], member[7], member[8]]
                    if member[5] == "done":
                        running.append(entry)
                await message.answer(f"""
زانیاری سەرجەم بەشداربووانی گرووپی تایبەت

سەرجەم بەردەوامبووەكان : {len(running)}

{date.today()}

""",parse_mode=types.ParseMode.MARKDOWN)
                def apply_formatting(sheet):
                    header_row = sheet[1]
                    for cell in header_row:
                        cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                def set_column_width(sheet, width):
                    for col in sheet.columns:
                        sheet.column_dimensions[col[0].column_letter].width = width
                def create_and_save_sheet(name, data, filename):
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title = name
                    for row in data:
                        sheet.append(row)
                    apply_formatting(sheet)
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                        for cell in row:
                            cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    sheet.sheet_view.rightToLeft = True
                    set_column_width(sheet, 18)
                    wb.save(filename)
                create_and_save_sheet("Running", running, "داتای بەردەوامبووان.xlsx")
                with open("داتای بەردەوامبووان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
            if message.text == "سەرجەم بەسەرچووەكان":
                header = ["ز", "ئایدی", "ڕۆژی بەشداری كردن", "ڕۆژی بەسەرچوون", "جۆری بەشداری كردن", "دۆخ", "نرخ", "جۆری پارەدان","#"]
                expired = [header]
                for i, member in enumerate(members, start=1):
                    entry = [i, member[1], member[2], member[3], member[4], member[5], member[6], member[7], member[8]]
                    if member[5] == "expired":
                        expired.append(entry)
                await message.answer(f"""
زانیاری سەرجەم بەشداربووانی گرووپی تایبەت

سەرجەم بەردەوامبووەكان : {len(expired)}

{date.today()}

""",parse_mode=types.ParseMode.MARKDOWN)
                def apply_formatting(sheet):
                    header_row = sheet[1]
                    for cell in header_row:
                        cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                def set_column_width(sheet, width):
                    for col in sheet.columns:
                        sheet.column_dimensions[col[0].column_letter].width = width
                def create_and_save_sheet(name, data, filename):
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title = name
                    for row in data:
                        sheet.append(row)
                    apply_formatting(sheet)
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                        for cell in row:
                            cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    sheet.sheet_view.rightToLeft = True
                    set_column_width(sheet, 18)
                    wb.save(filename)
                create_and_save_sheet("Running", expired, "داتای بەسەرچووەكان.xlsx")
                with open("داتای بەسەرچووەكان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
            if message.text == "سەرجەم چاوەڕوانی پارەدان":
                header = ["ز", "ئایدی", "ڕۆژی بەشداری كردن", "ڕۆژی بەسەرچوون", "جۆری بەشداری كردن", "دۆخ", "نرخ", "جۆری پارەدان","#"]
                pedding = [header]
                for i, member in enumerate(members, start=1):
                    entry = [i, member[1], member[2], member[3], member[4], member[5], member[6], member[7], member[8]]
                    if member[5] == "payment_pedding":
                        pedding.append(entry)
                await message.answer(f"""
زانیاری سەرجەم بەشداربووانی گرووپی تایبەت

سەرجەم بەسەرچووەكان : {len(pedding)}

{date.today()}

""",parse_mode=types.ParseMode.MARKDOWN)
                def apply_formatting(sheet):
                    header_row = sheet[1]
                    for cell in header_row:
                        cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                def set_column_width(sheet, width):
                    for col in sheet.columns:
                        sheet.column_dimensions[col[0].column_letter].width = width
                def create_and_save_sheet(name, data, filename):
                    wb = Workbook()
                    sheet = wb.active
                    sheet.title = name
                    for row in data:
                        sheet.append(row)
                    apply_formatting(sheet)
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                        for cell in row:
                            cell.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
                    sheet.sheet_view.rightToLeft = True
                    set_column_width(sheet, 18)
                    wb.save(filename)
                create_and_save_sheet("Running", pedding, "داتای بەسەرچووەكان.xlsx")
                with open("داتای بەسەرچووەكان.xlsx", 'rb') as document:
                    await self.bot.send_document(message.from_user.id, document)
        @self.dp.message_handler(lambda m:m.text == "بەشداربووانی بۆت")
        async def members_cmd(message:types.Message):
            members = db.userdata()
            await message.answer(f"""
سەرجەم بەشداربووەكان : {len(members)}

{date.today()}

""",parse_mode=types.ParseMode.MARKDOWN)
        @self.dp.message_handler(lambda m:m.text == "نامەی بەخێرهاتن")
        async def cmd_wlc_message(message: types.Message):
            welcome_message = config.get('welcomeMessage')
            await message.answer(welcome_message, reply_markup=keb.welcome_message_admin())

        @self.dp.callback_query_handler(lambda c: c.data.startswith('change_message'))
        async def process_callback_change_message(callback_query: types.CallbackQuery):
            await callback_query.message.answer('نامەی بەخێرهاتن بنووسە :')
            await ChannelsEdit.MESSAGE.set()


        @self.dp.message_handler(state=ChannelsEdit.MESSAGE)
        async def process_change_message(message: types.Message, state:FSMContext):
            new_message = message.text
            await state.update_data(message = new_message)
            await message.answer("ئایا دڵنیاییت لە گۆڕینی نامەی سەرەتا ؟",reply_markup=keb.get_confirm_of_create_account())
            await ChannelsEdit.CONFRIM.set()
        @self.dp.message_handler(lambda message: message.text in ['بەڵێ',"نەخێر"],state=ChannelsEdit.CONFRIM)
        async def check_status(message: types.Message,state:FSMContext):
            if message.text == "بەڵێ":
                data = await state.get_data()
                if data.get('photo'):
                    config['welcome_photo'] = data.get("photo")
                    write_config(config)
                    await message.answer("یەسەركەوتووی زیادكرا")
                    await self.bot.send_photo(message.from_user.id,photo=config.get("welcome_photo"),caption = config.get("welcomeMessage"),reply_markup=keb.welcome_message_admin())
                elif data.get("button"):
                    button_key = f"button"
                    button_value = {"title": f"{data.get('title')}", "description": f"{data.get('des')}"}
                    config["welcome_buttons"].append({button_key: button_value})
                    write_config(config)
                elif data.get("add_price"):
                    await message.answer("ماوە بنووسە بەم شێوازە => 1 مانگ ، 1 ساڵ")
                    await vip_signal_state.ADD_PRICE_TO_VIP_SIGNAL_TITLE.set()
                elif data.get("remove_admin_id"):
                    try:
                        db.update_column("users", "role", "member", "user_id", data.get("remove_admin_id"))
                        await message.answer("بەسەركەوتووی گۆڕدرا",reply_markup=keb.get_owner_keyboards(message.from_user.id))
                    except Exception as e:
                        message.answer(e)
                    await state.finish()
                else:
                    config['welcomeMessage'] = data.get("message")
                    write_config(config)
                    await message.answer("یەسەركەوتووی گۆڕدرا")
                    await message.answer(data.get("message"),reply_markup=keb.get_owner_keyboards(message.from_user.id))
                    await state.finish()
            if message.text == "نەخێر":
                await message.answer("هەڵوەشایەوە")
                await message.answer(data.get("message"),reply_markup=keb.welcome_message_admin())
            await state.finish()

        @self.dp.callback_query_handler(lambda c: c.data.startswith('photo_cmd'))
        async def process_callback_change_message(callback_query: types.CallbackQuery):
            if config.get("welcome_photo") == "Null":
                await callback_query.message.answer("هیچ وێنەیەك نیە",reply_markup=keb.photo_welcome_msg())
            else:

                await callback_query.message.answer("pas",reply_markup=keb.photo_welcome_msg(photo=True))

        @self.dp.callback_query_handler(lambda c: c.data.startswith('add_photo_wlc'))
        async def process_callback_add_photo(callback_query: types.CallbackQuery):
            await callback_query.message.answer("زۆر باشە ، تكایە وێنە بنێرە :",reply_markup=keb.back_confirm())
            await ChannelsEdit.PHOTO.set()

        @self.dp.message_handler(state=ChannelsEdit.PHOTO, content_types=types.ContentType.PHOTO)
        async def cmd_photo(message: types.Message, state: FSMContext):
            await state.update_data(photo=message.photo[-1].file_id)
            await message.answer("ئایا دڵنایی لە زیادكردنی وێنە؟", reply_markup=keb.get_confirm_of_create_account())
            await ChannelsEdit.CONFRIM.set()

        @self.dp.message_handler(state=ChannelsEdit.PHOTO, content_types=types.ContentType.TEXT)
        async def cmd_text(message: types.Message, state: FSMContext):
            await message.answer("هەڵوەشایەوە")
            await message.answer(data.get("message"), reply_markup=keb.welcome_message_admin())
            await state.finish()

        @self.dp.callback_query_handler(lambda c: c.data.startswith('inline_button'))
        async def process_callback_inline(callback_query: types.CallbackQuery):
            if not config["welcome_buttons"]:
                await callback_query.message.answer("هیچ دووگمەیەك نیە",reply_markup=keb.inline_welcome_msg())
        @self.dp.callback_query_handler(lambda c: c.data.startswith('add_inline_wlc'))
        async def process_callbackaddinline(callback_query: types.CallbackQuery):
            await callback_query.message.answer("تكایە ناوی دووگمە بنووسە :")
            await ChannelsEdit.INLINE.set()

        @self.dp.message_handler(state=ChannelsEdit.INLINE)
        async def process_name(message : types.Message,state:FSMContext):
            await state.update_data(title = message.text)
            await message.answer("تكایە دەقی دووگمە بنووسە :")
            await ChannelsEdit.INLINE_DES.set()

        @self.dp.message_handler(state=ChannelsEdit.INLINE_DES)
        async def process_name(message : types.Message,state:FSMContext):
            await state.update_data(des = message.text)
            await state.update_data(button = "True")
            await message.answer("ئایا دڵنایی لە زیادكردنی ؟", reply_markup=keb.get_confirm_of_create_account())
            await ChannelsEdit.CONFRIM.set()

        @self.dp.message_handler(lambda m:m.text == "نرخی گرووپی سیگناڵ")
        async def price_vip_signal(message:types.Message):
            print("sss")
            vip_signal_prices = config.get("vip_signal_price")
            if "vip_signal_price" in config and (config["vip_signal_price"] == [] or config["vip_signal_price"] is None):
                print("vip_signal_price is null")
            else:
                # print("8hs")
                title = []
                for payment_info in vip_signal_prices:
                    payment = payment_info.get("payment", {})
                    title.append(payment.get("name", "N/A"))
                    name = payment.get("name", "N/A")
                    price = payment.get("price", "N/A")

                    print("Name:", name)
                    print("Price:", price)
                await message.answer("ss",reply_markup=anal.vip_signals_price(payment=title))

        @self.dp.callback_query_handler(lambda c:c.data == "add_payment_price")
        async def add_price_v_signal(call : types.CallbackQuery,state=FSMContext):
            await call.message.answer("ئایا دڵنایی لە زیادكردنی نرخی نوێ", reply_markup=keb.get_confirm_of_create_account())
            await state.update_data(add_price = "True")
            await ChannelsEdit.CONFRIM.set()
        @self.dp.message_handler(state=vip_signal_state.ADD_PRICE_TO_VIP_SIGNAL_TITLE)
        async def ADD_PRICE_TO_VIP_SIGNAL_PRICE_cmd(message:types.Message):

            await message.answer("زۆر باشە ، تكایە نرخ بنووسە بەم شێوازە => $45 , $30")
        @self.dp.message_handler(lambda m:m.text == "ئەدمینەكان")
        async def admins_cmd(message:types.Message):
            await message.answer(f"""
{message.from_user.full_name}

ئەمە سەرجەم ئەدمینەكانن

بۆ دەركردن گرتە بكە لە ناوەكە

بۆ زیادكردنی ئەدمینی نوێ گرتە بكە لە زیادكردنی ئەدمین
""",reply_markup=anal.get_admin_anal(),parse_mode=types.ParseMode.MARKDOWN)
        @self.dp.callback_query_handler(lambda c:c.data.startswith("admin_cl__"))
        async def add_price_v_signal(call : types.CallbackQuery,state=FSMContext):
            user_id = call.data.split("__")[1]
            get_detail_of_user = db.userdata(user_id=user_id)
            await call.message.answer(f"""

ناو : {get_detail_of_user[1]}
نازناو : @{get_detail_of_user[2]}
ڕۆڵ : {get_detail_of_user[4]}
بەروار : {get_detail_of_user[5]}
""",reply_markup=anal.remove_admin(user_id))
        @self.dp.callback_query_handler(lambda c:c.data.startswith("delete_admin__"))
        async def remove_admin_confirmation(call : types.CallbackQuery,state=FSMContext):
            user_id = call.data.split("__")[1]
            await state.update_data(remove_admin_id = user_id)
            await call.message.answer("ئايا دڵنیایی لە ڕەشكردنەوەی ؟",reply_markup=keb.get_confirm_of_create_account())
            await ChannelsEdit.CONFRIM.set()